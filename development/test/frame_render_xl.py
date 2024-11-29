from   ezpath  import *
from   ezhdl   import *
from   vic_pkg import *

import bus_logger as bl
import numpy as np
import openpyxl as xl
from   openpyxl.styles import PatternFill, Alignment, Font

MAX_HRES = 504
MAX_VRES = 312

class FrameRenderXl(Entity):
	def __init__(self):
		self.i_clk  = Input(Wire())
		self.i_rst  = Input(Wire())
		self.i_push = Input(Wire())
		self.i_lstr = Input(Wire())
		self.i_lend = Input(Wire())
		self.i_fstr = Input(Wire())
		self.i_colr = Input(t_vic_colr)

		self.xpos  = 0
		self.ypos  = 0

		self.wb = xl.Workbook()
		self.sh = self.wb.active
		self.sh.title = "Frame Dump"

		self.color = [
			"010101", # black
			"ffffff", # white
			"9f4e44", # red
			"6abfc6", # cyan
			"a057a3", # purple
			"5cab5e", # green
			"50459b", # blue
			"c9d487", # yellow
			"a1683c", # orange
			"6d5412", # brown
			"cb7e75", # pink
			"626262", # dark grey
			"898989", # grey
			"9ae29b", # light green
			"887ecb", # light blue
			"adadad"  # light grey
		]

		self.color_txt = [
			"ffffff", # black
			"010101", # white
			"ffffff", # red
			"010101", # cyan
			"ffffff", # purple
			"ffffff", # green
			"ffffff", # blue
			"010101", # yellow
			"010101", # orange
			"ffffff", # brown
			"010101", # pink
			"ffffff", # dark grey
			"010101", # grey
			"010101", # light green
			"010101", # light blue
			"010101"  # light grey
		]

	def _run(self):
		if self.i_clk.negedge():
			if self.i_push.now == 1:
				if(self.xpos > MAX_HRES - 1):
					self.xpos = MAX_HRES - 1

				if(self.ypos > MAX_VRES - 1):
					self.ypos = MAX_VRES - 1

				if self.i_lstr.now == 1:
					self.xpos = 0
					if self.i_fstr.now == 1:
						self.ypos = 0

				color = self.color[self.i_colr.now]
				color_txt = self.color_txt[self.i_colr.now]
				font = Font(name="Consolas", color=color_txt, bold=True)
				fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
				alignment = Alignment(horizontal="left", vertical="top")

				self.sh.cell(row=self.ypos+1, column=self.xpos+1, value= "     \n" + "\n".join(bl.log))
				self.sh.cell(row=self.ypos+1, column=self.xpos+1).fill = fill
				self.sh.cell(row=self.ypos+1, column=self.xpos+1).font = font
				self.sh.cell(row=self.ypos+1, column=self.xpos+1).alignment = alignment
				bl.clear()

				self.xpos += 1
				if self.i_lend.now == 1:
					self.ypos += 1

			if self.i_clk.now:
				pass


	def save(self, path):
		cell_size = 12
		print("Saving excel dump")

		# Adjusting cell width
		for col in self.sh.columns:
			for cell in col:
				self.sh.column_dimensions[cell.column_letter].width = cell_size / 5  # Rough estimation of character width

		# Adjusting cell height
		for row in self.sh.rows:
			for cell in row:
				self.sh.row_dimensions[cell.row].height = cell_size  # Set height in pixels

		print("Finalizing")
		self.wb.save(get_abs_path(path))
